"""
backend/app/routes/bulk_router.py
Full Bulk Operations API — enterprise-grade batch actions across profiles

Endpoints:
  POST /api/v1/bulk/profiles/action     — Approve / Certify / Return / Assign / Reassign / Export
  POST /api/v1/bulk/exceptions/resolve  — Bulk resolve exceptions with root-cause
  POST /api/v1/bulk/exceptions/assign   — Bulk assign exceptions to a user
  POST /api/v1/bulk/profiles/export     — Export selected profiles to Excel/CSV
  GET  /api/v1/bulk/profiles            — Paginated + filtered profile list (for the UI table)
  GET  /api/v1/bulk/users               — User list (for assign dropdowns)
  GET  /api/v1/bulk/summary             — Counts by lifecycle state (for filter badges)
"""
import io
import logging
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import text

from ..database import get_db
from ..rbac.roles import ADMIN, PREPARER, APPROVER, CERTIFIER
from ..rbac.dependencies import role_required

log = logging.getLogger(__name__)

router = APIRouter(prefix="/api/v1/bulk", tags=["bulk-operations"])

_ALL      = [ADMIN, PREPARER, APPROVER, CERTIFIER]
_ACTIONERS = [ADMIN, APPROVER]
_CERTIFIERS = [ADMIN, CERTIFIER]

# ── Valid lifecycle actions ────────────────────────────────────────────────────
PROFILE_ACTIONS = {
    "APPROVE":   {"to": "APPROVED",    "roles": [ADMIN, APPROVER],  "from": {"SUBMITTED", "REVIEWED", "UNDER_REVIEW"}},
    "CERTIFY":   {"to": "CERTIFIED",   "roles": [ADMIN, CERTIFIER], "from": {"APPROVED"}},
    "RETURN":    {"to": "PREPARED",    "roles": [ADMIN, APPROVER],  "from": {"SUBMITTED", "REVIEWED", "UNDER_REVIEW", "APPROVED"}},
    "CLOSE":     {"to": "CLOSED",      "roles": [ADMIN],            "from": {"CERTIFIED"}},
    "REOPEN":    {"to": "OPEN",        "roles": [ADMIN],            "from": None},       # any state
    "ASSIGN":    {"to": None,          "roles": [ADMIN, APPROVER, PREPARER], "from": None},
}

ROOT_CAUSES = {
    "TIMING_DIFFERENCE", "DUPLICATE", "SYSTEM_ERROR",
    "MANUAL_ERROR", "FX_ROUNDING", "POLICY_GAP",
    "DATA_QUALITY", "AWAITING_CONFIRMATION", "OTHER",
}


# ── Schemas ───────────────────────────────────────────────────────────────────

class ProfileBulkActionRequest(BaseModel):
    profile_ids:    List[int]
    action:         str          # APPROVE | CERTIFY | RETURN | CLOSE | REOPEN | ASSIGN
    comments:       Optional[str] = None
    assign_user_id: Optional[int] = None   # for ASSIGN action

class ExceptionBulkResolveRequest(BaseModel):
    exception_ids: List[int]
    root_cause:    str        # from ROOT_CAUSES enum
    resolution_notes: Optional[str] = None
    status:        str = "RESOLVED"

class ExceptionBulkAssignRequest(BaseModel):
    exception_ids: List[int]
    assign_user_id: int
    comments:       Optional[str] = None

class BulkExportRequest(BaseModel):
    profile_ids: List[int]
    format:      str = "xlsx"   # xlsx | csv


# ── Helpers ───────────────────────────────────────────────────────────────────

def _profile_row(p, db: Session) -> dict:
    """Lightweight profile dict for the bulk table."""
    from ..models.models import User, ReconciliationBalance, ExceptionQueueRecord, MatchGroup
    preparer = db.query(User).filter(User.id == p.assigned_preparer).first() if getattr(p, "assigned_preparer", None) else None
    approver = db.query(User).filter(User.id == p.assigned_reviewer).first() if getattr(p, "assigned_reviewer", None) else None

    open_exceptions = db.query(ExceptionQueueRecord).join(
        MatchGroup, MatchGroup.id == ExceptionQueueRecord.match_group_id
    ).filter(
        MatchGroup.profile_id == p.id,
        ExceptionQueueRecord.status == "OPEN",
    ).count()

    last_balance = db.query(ReconciliationBalance).filter(
        ReconciliationBalance.profile_id == p.id
    ).order_by(ReconciliationBalance.id.desc()).first()

    return {
        "id":                   p.id,
        "name":                 p.name,
        "account_number":       getattr(p, "account_number", None),
        "period":               getattr(p, "period", None),
        "reconciliation_type":  getattr(p, "reconciliation_type", None),
        "lifecycle_state":      getattr(p, "lifecycle_state", "OPEN"),
        "risk_classification":  getattr(p, "risk_classification", "LOW"),
        "frequency":            getattr(p, "frequency", None),
        "assigned_preparer_id": getattr(p, "assigned_preparer", None),
        "assigned_preparer":    preparer.email if preparer else None,
        "assigned_reviewer_id": getattr(p, "assigned_reviewer", None),
        "assigned_reviewer":    approver.email if approver else None,
        "open_exceptions":      open_exceptions,
        "balance_status":       last_balance.status if last_balance else None,
        "variance":             float(last_balance.variance_amount or 0) if last_balance else None,
        "created_at":           p.created_at.isoformat() if p.created_at else None,
        "updated_at":           p.updated_at.isoformat() if getattr(p, "updated_at", None) else None,
    }


# ── GET /profiles — paginated profile list ────────────────────────────────────

@router.get("/profiles")
def list_profiles_bulk(
    search:          Optional[str] = Query(None),
    lifecycle_state: Optional[str] = Query(None),
    risk:            Optional[str] = Query(None),
    recon_type:      Optional[str] = Query(None),
    period:          Optional[str] = Query(None),
    has_exceptions:  Optional[bool] = Query(None),
    sort_by:         str  = Query("id"),
    sort_dir:        str  = Query("desc"),
    limit:           int  = Query(200, ge=1, le=1000),
    offset:          int  = Query(0, ge=0),
    db: Session = Depends(get_db),
    current_user=Depends(role_required(_ALL)),
):
    from ..models.models import ReconciliationProfile as RP

    q = db.query(RP)

    if lifecycle_state and lifecycle_state.upper() != "ALL":
        states = [s.strip().upper() for s in lifecycle_state.split(",")]
        q = q.filter(RP.lifecycle_state.in_(states))
    if risk and risk.upper() != "ALL":
        q = q.filter(RP.risk_classification == risk.upper())
    if recon_type:
        q = q.filter(RP.reconciliation_type == recon_type)
    if period:
        q = q.filter(RP.period == period)
    if search:
        like = f"%{search}%"
        q = q.filter(RP.name.ilike(like) | RP.account_number.ilike(like))

    total = q.count()

    # Sort
    col = getattr(RP, sort_by, RP.id)
    q = q.order_by(col.desc() if sort_dir == "desc" else col.asc())
    profiles = q.offset(offset).limit(limit).all()

    rows = [_profile_row(p, db) for p in profiles]

    # Filter by has_exceptions after fetching (lightweight)
    if has_exceptions is True:
        rows = [r for r in rows if r["open_exceptions"] > 0]
    elif has_exceptions is False:
        rows = [r for r in rows if r["open_exceptions"] == 0]

    return {"profiles": rows, "total": total, "offset": offset, "limit": limit}


# ── GET /summary — counts by state ────────────────────────────────────────────

@router.get("/summary")
def bulk_summary(
    db: Session = Depends(get_db),
    current_user=Depends(role_required(_ALL)),
):
    from ..models.models import ReconciliationProfile as RP

    rows = db.execute(text(
        "SELECT lifecycle_state, COUNT(*) as cnt "
        "FROM reconciliation_profiles GROUP BY lifecycle_state"
    )).fetchall()
    by_state = {r.lifecycle_state or "OPEN": int(r.cnt) for r in rows}

    risk_rows = db.execute(text(
        "SELECT risk_classification, COUNT(*) as cnt "
        "FROM reconciliation_profiles GROUP BY risk_classification"
    )).fetchall()
    by_risk = {r.risk_classification or "UNKNOWN": int(r.cnt) for r in risk_rows}

    open_exc = db.execute(text(
        "SELECT COUNT(*) FROM exception_queue_records WHERE status = 'OPEN'"
    )).scalar()

    total = db.query(RP).count()

    return {
        "total":        total,
        "by_state":     by_state,
        "by_risk":      by_risk,
        "open_exceptions": int(open_exc or 0),
    }


# ── GET /users — user list for assign dropdowns ───────────────────────────────

@router.get("/users")
def list_users(
    db: Session = Depends(get_db),
    current_user=Depends(role_required(_ALL)),
):
    from ..models.models import User
    users = db.query(User).filter(User.is_active == True).all()
    return [
        {"id": u.id, "email": u.email, "role": u.role, "full_name": getattr(u, "full_name", u.email)}
        for u in users
    ]


# ── POST /profiles/action — core bulk action ──────────────────────────────────

@router.post("/profiles/action")
def bulk_profile_action(
    payload: ProfileBulkActionRequest,
    db: Session = Depends(get_db),
    current_user=Depends(role_required(_ALL)),
):
    from ..models.models import ReconciliationProfile as RP, ReconciliationBalance as RB, User
    from ..services import audit_service

    action = payload.action.upper().strip()
    if action not in PROFILE_ACTIONS:
        raise HTTPException(status_code=400, detail=f"Unknown action '{action}'. Valid: {list(PROFILE_ACTIONS)}")

    meta = PROFILE_ACTIONS[action]

    # Role check
    if current_user.role not in meta["roles"]:
        raise HTTPException(status_code=403, detail=f"Role '{current_user.role}' cannot perform '{action}'")

    if action == "ASSIGN" and not payload.assign_user_id:
        raise HTTPException(status_code=400, detail="assign_user_id is required for ASSIGN action")

    if not payload.profile_ids:
        raise HTTPException(status_code=400, detail="No profile IDs provided")

    profiles = db.query(RP).filter(RP.id.in_(payload.profile_ids)).all()
    if not profiles:
        raise HTTPException(status_code=404, detail="No matching profiles found")

    now = datetime.utcnow()
    results = {"success": [], "skipped": [], "errors": []}

    for p in profiles:
        try:
            current_state = (getattr(p, "lifecycle_state", "OPEN") or "OPEN").upper()

            if action == "ASSIGN":
                # Assign reviewer/preparer depending on user role
                target_user = db.query(User).filter(User.id == payload.assign_user_id).first()
                if not target_user:
                    results["errors"].append({"id": p.id, "reason": "Target user not found"})
                    continue

                if target_user.role in (PREPARER,):
                    p.assigned_preparer = payload.assign_user_id
                else:
                    p.assigned_reviewer = payload.assign_user_id

                audit_service.log_action(db, "BULK_ASSIGN", user_id=current_user.id,
                    entity_type="profile", entity_id=p.id,
                    metadata={"assigned_to": payload.assign_user_id, "comments": payload.comments})
                results["success"].append(p.id)
                continue

            # State transition actions
            allowed_from = meta["from"]
            if allowed_from and current_state not in allowed_from:
                results["skipped"].append({
                    "id": p.id,
                    "reason": f"State '{current_state}' not eligible for '{action}' (allowed: {sorted(allowed_from)})"
                })
                continue

            target_state = meta["to"]
            p.lifecycle_state = target_state

            # Also update the latest ReconciliationBalance status to match
            latest_balance = db.query(RB).filter(
                RB.profile_id == p.id
            ).order_by(RB.id.desc()).first()

            if latest_balance:
                balance_state_map = {
                    "APPROVED":  "APPROVED",
                    "CERTIFIED": "CERTIFIED",
                    "PREPARED":  "DRAFT",
                    "CLOSED":    "CERTIFIED",
                    "OPEN":      "DRAFT",
                }
                new_bal_state = balance_state_map.get(target_state)
                if new_bal_state:
                    latest_balance.status = new_bal_state
                    if target_state == "APPROVED":
                        latest_balance.approved_by = current_user.id
                        latest_balance.approved_at = now
                    elif target_state == "CERTIFIED":
                        latest_balance.certified_by = current_user.id
                        latest_balance.certified_at = now

            audit_service.log_action(db, f"BULK_{action}", user_id=current_user.id,
                entity_type="profile", entity_id=p.id,
                metadata={"from_state": current_state, "to_state": target_state, "comments": payload.comments})
            results["success"].append(p.id)

        except Exception as e:
            log.warning(f"[bulk_action] profile {p.id} error: {e}")
            results["errors"].append({"id": p.id, "reason": str(e)})

    db.commit()

    return {
        "action":    action,
        "processed": len(profiles),
        "success":   len(results["success"]),
        "skipped":   len(results["skipped"]),
        "errors":    len(results["errors"]),
        "detail":    results,
    }


# ── POST /exceptions/resolve — bulk resolve exceptions ────────────────────────

@router.post("/exceptions/resolve")
def bulk_resolve_exceptions(
    payload: ExceptionBulkResolveRequest,
    db: Session = Depends(get_db),
    current_user=Depends(role_required(_ACTIONERS)),
):
    from ..models.models import ExceptionQueueRecord as EQR
    from ..services import audit_service

    root_cause = payload.root_cause.upper().strip()
    if root_cause not in ROOT_CAUSES:
        raise HTTPException(status_code=400, detail=f"Invalid root_cause. Valid: {sorted(ROOT_CAUSES)}")

    target_status = payload.status.upper()
    if target_status not in ("RESOLVED", "IN_PROGRESS", "ESCALATED"):
        raise HTTPException(status_code=400, detail="status must be RESOLVED, IN_PROGRESS, or ESCALATED")

    if not payload.exception_ids:
        raise HTTPException(status_code=400, detail="No exception IDs provided")

    excs = db.query(EQR).filter(EQR.id.in_(payload.exception_ids)).all()
    now  = datetime.utcnow()
    resolved = 0

    for exc in excs:
        try:
            exc.status           = target_status
            exc.classification   = root_cause
            exc.resolution_notes = payload.resolution_notes
            if target_status == "RESOLVED":
                exc.resolved_at = now

            audit_service.log_action(db, "BULK_EXCEPTION_RESOLVED", user_id=current_user.id,
                entity_type="exception", entity_id=exc.id,
                metadata={"root_cause": root_cause, "notes": payload.resolution_notes, "status": target_status})
            resolved += 1
        except Exception as e:
            log.warning(f"[bulk_resolve] exception {exc.id} error: {e}")

    db.commit()
    return {"resolved": resolved, "total_requested": len(payload.exception_ids)}


# ── POST /exceptions/assign — bulk assign exceptions ─────────────────────────

@router.post("/exceptions/assign")
def bulk_assign_exceptions(
    payload: ExceptionBulkAssignRequest,
    db: Session = Depends(get_db),
    current_user=Depends(role_required(_ACTIONERS)),
):
    from ..models.models import ExceptionQueueRecord as EQR
    from ..services import audit_service

    if not payload.exception_ids:
        raise HTTPException(status_code=400, detail="No exception IDs provided")

    excs = db.query(EQR).filter(EQR.id.in_(payload.exception_ids)).all()
    assigned = 0
    for exc in excs:
        exc.assigned_to = payload.assign_user_id
        exc.status      = "IN_PROGRESS"
        if payload.comments:
            exc.comments = payload.comments
        audit_service.log_action(db, "BULK_EXCEPTION_ASSIGNED", user_id=current_user.id,
            entity_type="exception", entity_id=exc.id,
            metadata={"assigned_to": payload.assign_user_id})
        assigned += 1

    db.commit()
    return {"assigned": assigned}


# ── POST /profiles/export — export selected profiles ─────────────────────────

@router.post("/profiles/export")
def export_profiles(
    payload: BulkExportRequest,
    db: Session = Depends(get_db),
    current_user=Depends(role_required(_ALL)),
):
    from ..models.models import ReconciliationProfile as RP

    profiles = db.query(RP).filter(RP.id.in_(payload.profile_ids)).all() if payload.profile_ids else db.query(RP).all()
    rows = [_profile_row(p, db) for p in profiles]

    fmt = payload.format.lower()

    if fmt == "csv":
        import csv
        output = io.StringIO()
        if rows:
            writer = csv.DictWriter(output, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=profiles_{datetime.utcnow().strftime('%Y%m%d')}.csv"},
        )
    else:
        # Excel via openpyxl
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Profiles"

        HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
        HEADER_FONT = Font(bold=True, color="FFE600", size=10)
        headers = ["ID", "Name", "Account", "Period", "Type", "State", "Risk",
                   "Preparer", "Reviewer", "Open Exceptions", "Variance", "Balance Status", "Created"]
        field_map = ["id", "name", "account_number", "period", "reconciliation_type",
                     "lifecycle_state", "risk_classification", "assigned_preparer", "assigned_reviewer",
                     "open_exceptions", "variance", "balance_status", "created_at"]

        for col_idx, hdr in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=hdr)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        STATE_COLORS = {
            "CERTIFIED": "00C891", "APPROVED": "2d8cf0", "SUBMITTED": "f59e0b",
            "PREPARED": "94A3B8", "OPEN": "6B7280", "CLOSED": "374151",
        }
        RISK_COLORS = {"HIGH": "FF4D4D", "MEDIUM": "FFE600", "LOW": "00C891"}

        for row_idx, r in enumerate(rows, 2):
            for col_idx, field in enumerate(field_map, 1):
                val = r.get(field)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = Font(size=10)
                # Color-code state and risk
                if field == "lifecycle_state" and val in STATE_COLORS:
                    cell.fill = PatternFill("solid", fgColor=STATE_COLORS[val])
                    cell.font = Font(bold=True, color="FFFFFF", size=10)
                elif field == "risk_classification" and val in RISK_COLORS:
                    cell.fill = PatternFill("solid", fgColor=RISK_COLORS[val])
                    cell.font = Font(bold=True, color="FFFFFF", size=10)

        # Auto-size columns
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        # Freeze header row
        ws.freeze_panes = "A2"

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=profiles_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"},
        )
