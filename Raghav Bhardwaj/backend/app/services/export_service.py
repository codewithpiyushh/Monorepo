from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, Iterable, List, Optional

import pandas as pd
from openpyxl.styles import Font, PatternFill


DEFAULT_EXPORT_COLUMNS = [
    "transaction_id",
    "source_amount",
    "target_amount",
    "difference",
    "status",
]


def _to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        text = str(value).replace(",", "").strip()
        if text == "":
            return None
        return float(text)
    except (ValueError, TypeError):
        return None


def _extract_amount(payload: Optional[dict]) -> Optional[float]:
    if not payload:
        return None

    preferred_keys = ("amount", "amt", "value", "balance", "debit", "credit")
    normalized = {str(k).strip().lower(): v for k, v in payload.items()}

    for key in preferred_keys:
        if key in normalized:
            parsed = _to_float(normalized[key])
            if parsed is not None:
                return parsed

    for value in payload.values():
        parsed = _to_float(value)
        if parsed is not None:
            return parsed
    return None


def format_reconciliation_data(records: Iterable[Any]) -> List[Dict[str, Any]]:
    output: List[Dict[str, Any]] = []
    for record in records:
        source_data = json.loads(record.source_data) if getattr(record, "source_data", None) else {}
        target_data = json.loads(record.target_data) if getattr(record, "target_data", None) else {}
        source_amount = _extract_amount(source_data)
        target_amount = _extract_amount(target_data)
        difference = None
        if source_amount is not None and target_amount is not None:
            difference = source_amount - target_amount

        output.append(
            {
                "transaction_id": getattr(record, "id", None),
                "source_amount": source_amount,
                "target_amount": target_amount,
                "difference": difference,
                "status": getattr(record, "match_status", None),
            }
        )
    return output


def calculate_summary(matched: List[Dict[str, Any]], unmatched: List[Dict[str, Any]]) -> Dict[str, Any]:
    matched_count = len(matched)
    unmatched_count = len(unmatched)
    total_records = matched_count + unmatched_count
    match_percentage = round((matched_count / total_records) * 100, 2) if total_records else 0.0
    return {
        "total_records": total_records,
        "matched_count": matched_count,
        "unmatched_count": unmatched_count,
        "match_percentage": match_percentage,
    }


def _autosize_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_len = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_len + 2, 10), 40)


def _format_sheet(worksheet, numeric_cols: Iterable[str]) -> None:
    worksheet.freeze_panes = "A2"
    header_font = Font(bold=True)
    header_row = next(worksheet.iter_rows(min_row=1, max_row=1))
    for cell in header_row:
        cell.font = header_font

    header_to_index = {cell.value: idx + 1 for idx, cell in enumerate(header_row) if cell.value}
    for col_name in numeric_cols:
        col_idx = header_to_index.get(col_name)
        if not col_idx:
            continue
        for row in worksheet.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    cell.number_format = "#,##0.00"

    _autosize_columns(worksheet)


def _apply_unmatched_highlight(worksheet) -> None:
    red_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.fill = red_fill


def generate_excel_report(data: dict) -> BytesIO:
    matched_data = data.get("matched", [])
    unmatched_data = data.get("unmatched", [])
    unit_summary = data.get("unit_summary", [])
    transactions_data = data.get("transactions", [])
    summary_data = data.get("summary", {})
    metadata = data.get("metadata", {})
    columns = data.get("columns") or DEFAULT_EXPORT_COLUMNS

    matched_df = pd.DataFrame(matched_data)
    unmatched_df = pd.DataFrame(unmatched_data)
    unit_summary_df = pd.DataFrame(unit_summary)
    transactions_df = pd.DataFrame(transactions_data)
    summary_df = pd.DataFrame([summary_data]) if summary_data else pd.DataFrame([{}])
    metadata_df = pd.DataFrame([metadata]) if metadata else pd.DataFrame([{}])

    if not matched_df.empty:
        matched_df = matched_df.reindex(columns=columns)
    else:
        matched_df = pd.DataFrame(columns=columns)

    if not unmatched_df.empty:
        unmatched_df = unmatched_df.reindex(columns=columns)
    else:
        unmatched_df = pd.DataFrame(columns=columns)

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if not unit_summary_df.empty or not transactions_df.empty:
            if unit_summary_df.empty:
                unit_summary_df = pd.DataFrame(columns=["entity", "account", "status", "total_transactions", "matched_count", "partial_count", "unmatched_count"])
            if transactions_df.empty:
                transactions_df = pd.DataFrame(columns=columns)
            unit_summary_df.to_excel(writer, index=False, sheet_name="Unit_Summary")
            transactions_df.to_excel(writer, index=False, sheet_name="Transactions")
        else:
            matched_df.to_excel(writer, index=False, sheet_name="Matched")
            unmatched_df.to_excel(writer, index=False, sheet_name="Unmatched")
        summary_df.to_excel(writer, index=False, sheet_name="Run_Summary")
        metadata_df.to_excel(writer, index=False, sheet_name="Metadata")

        workbook = writer.book
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            _format_sheet(worksheet, numeric_cols=("source_amount", "target_amount", "difference", "match_percentage"))
        if "Unmatched" in workbook.sheetnames:
            _apply_unmatched_highlight(workbook["Unmatched"])

    output.seek(0)
    return output


def generate_csv_report(rows: List[Dict[str, Any]], columns: Optional[List[str]] = None) -> BytesIO:
    selected_columns = columns or DEFAULT_EXPORT_COLUMNS
    frame = pd.DataFrame(rows)
    if frame.empty:
        frame = pd.DataFrame(columns=selected_columns)
    else:
        frame = frame.reindex(columns=selected_columns)

    output = BytesIO()
    csv_bytes = frame.to_csv(index=False).encode("utf-8")
    output.write(csv_bytes)
    output.seek(0)
    return output


def build_filename(prefix: str, entity_id: int, extension: str) -> str:
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    return f"{prefix}_{entity_id}_{timestamp}.{extension}"
